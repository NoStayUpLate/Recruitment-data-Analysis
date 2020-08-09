from openpyxl import load_workbook
from 毕业设计.Data_handle import DataHandle
import xlrd

wb = load_workbook("./ExcelFile/职位要求1.xlsx")
ws = wb['Sheet1']

data = xlrd.open_workbook("./ExcelFile/职位要求1.xlsx",encoding_override='utf-8')
tabel = data.sheet_by_index(0)

for i in range(1,1307):
    SkillRequirements = tabel.row_values(i)[1]
    # print(SkillRequirements)
    """
    任职要求
    岗位要求
    技术要求
    任职资格
    职位要求
    职责要求
    知识技能
    任职条件
    工作要求
    我们对你的要求
    """
    # if "任职要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("任职要求")[1]
    # if "岗位要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("岗位要求")[1]
    # if "技术要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("技术要求")[1]
    # if "任职资格" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("任职资格")[1]
    # if "职位要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("职位要求")[1]
    # if "职责要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("职责要求")[1]
    # if "知识技能" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("知识技能")[1]
    # if "任职条件" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("任职条件")[1]
    # if "工作要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("工作要求")[1]
    # if "我们对你的要求" in SkillRequirements:
    #     SkillRequirements = SkillRequirements.split("我们对你的要求")[1]
    SkillRequirements = SkillRequirements.replace(" ","")

    ws.cell(row=i+1,column=2).value = SkillRequirements

    wb.save("./ExcelFile/职位要求1.xlsx")
    print("*-|-*")

print("结束！！！！！！！！！！！！！！！！！！！！！！！！！！！")
