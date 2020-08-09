import os
import random
import time
import redis
import requests
from lxml import etree
from openpyxl import load_workbook  # 导入 xlsx 表格操作包
from 毕业设计.Data_handle import DataHandle

# 城市设置
city = '%E4%B8%8A%E6%B5%B7'  # 上海  更换此值可以跟换其他城市的 数据
# 百度搜索 城市名  复制 url j链接后面的 wd=武汉
# 例子: https://www.baidu.com/s?wd=武汉
# 复制该链接到浏览器 回车  然后再复浏览器上的运行后的链接 粘贴到这里 即可得到城市的编码


# TODO 此处为新加的 读取职位的详细说明的代码
def get_detail(url):
    resp = requests.get(url,headers=headers)
    job_detail = etree.HTML(resp.text).xpath('//*[@id="job_djjjjjetail"]/dd[2]/div/text()')
    detail=''.join([i.strip() for i in job_detail])
    return detail

red = redis.Redis('127.0.0.1', 6379, decode_responses=True, db=0)

user_agent_list = [
    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11 ',
    'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1',
    'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
]
headers = {
    'user-agent': random.choice(user_agent_list),
}
cookies = {
    '_ga': 'GA1.2.498334904.1585298707',
    'LGUID': '20200327164507-be650f4a-d496-4d32-8717-287f8817bb2f',
    '_gid': 'GA1.2.1147502390.1588771320',
    'index_location_city': city,
    'WEBTJ-ID': '20200508233546-171f4eda4097e-0905cd590493ba-d373666-1327104-171f4eda40a8ce',
    'PRE_UTM': '',
    'PRE_LAND': 'https%3A%2F%2Fwww.lagou.com%2F',
    'LGSID': '20200508233545-ac3a5941-54a5-4b0a-ab89-a5d4962cb973',
    'PRE_HOST': 'www.baidu.com',
    'PRE_SITE': 'https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3Di0u2aM3bQ%5Fue65-pVY-8f%5FexgsxRdgmnH1NWUjULfcS%26wd%3D%26eqid%3Dd6fe9b3d00015a2f000000055eb57c46',
    '_gat': '1',
    'TG-TRACK-CODE': 'index_navigation',
    'LGRID': '20200508235025-17a32b42-147c-4c8a-a077-a454d3c08709',
}


with open('search_str.txt', mode='r', encoding='utf-8') as f:  # 读取 search_str.txt文件 里面存储的是需要搜索的职位关键词
    search_list = [i.strip() for i in f.readlines()]   # 把所有关键词保存到列表

def get_search_word():
    red_search_word = red.get('search_word')  # 读取redis 中的 查询关键词 (类似下载的断点续传)
    if red_search_word:
        red.set('search_word', search_list[search_list.index(red_search_word) + 1])
        yield search_list[search_list.index(red_search_word) + 1]
    else:
        yield 'Java'

def third_request(row, result, showId, search_word):
    wb = load_workbook('ExcelFile/拉钩-上海.xlsx')
    ws = wb['Sheet1']  # 指定 word表格 sheet 页面
    for i in result:
        row += 1
        url = 'https://www.lagou.com/jobs/%s.html?show=%s' % (i['positionId'], showId)
        # TODO 取消下面一行的注释 以及第95行的注释 即可把详情保存到表格中
        detail = get_detail(url)
        # print(detail)
        ws['A' + str(row)] = search_word  # 查询词
        ws['B' + str(row)] = i['positionName']  # 标题
        ws['C' + str(row)] = i['salary']  # 薪资范围
        Salary_min,Salary_max = DataHandle().DealSalary(i['salary'])
        ws['D' + str(row)] = i['companyFullName']  # 公司全称
        ws['E' + str(row)] = DataHandle().DealIndustryNull(i['industryField'])  # 公司类型
        ws['F' + str(row)] = i['companySize']  # 公司规模
        Company_size_Min,Company_size_Max = DataHandle().DealCompanySize(i['companySize'])
        ws['G' + str(row)] = DataHandle().DealPublishData(i['createTime'])  # 发布时间
        ws['H' + str(row)] = i['city']  # 城市
        ws['I' + str(row)] = DataHandle().DealAreaNull(i['district'])  # 区
        ws['J' + str(row)] = i['workYear']  # 年限要求
        Working_Seniority_Requirements_Min, Working_Seniority_Requirements_Max = DataHandle().DealWorkingSeniorityRequirements(i['workYear'])
        ws['K' + str(row)] = i['education']  # 学历要求
        ws['L' + str(row)] = i['positionAdvantage']  # 职位诱惑
        ws['M' + str(row)] = i['financeStage']  # 融资状态
        ws['N' + str(row)] = detail  #url  # 详情url
        # TODO 取消下面一行的注释 以及第77行的注释 即可把详情保存到表格中
        # ws['N' + str(row)] = detail

        ws['O' + str(row)] = Salary_min
        ws['P' + str(row)] = Salary_max
        ws['Q' + str(row)] = Company_size_Min
        ws['R' + str(row)] = Company_size_Max
        ws['S' + str(row)] = Working_Seniority_Requirements_Min
        ws['T' + str(row)] = Working_Seniority_Requirements_Max

    wb.save('拉钩-上海.xlsx')  # 保存到word表格
    wb.close()  # 关闭word表格
    print('保存成功')
    return row

def second_request(cookie, max_page_number, search_word):
    print('查询词:',search_word,'有', max_page_number,'页 可见数据')
    headers = {
        'Origin': 'https://www.lagou.com',
        'X-Anit-Forge-Code': '0',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'User-Agent': random.choice(user_agent_list),
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Referer': 'https://www.lagou.com/jobs/list_?px=new&city=%s'%city,
        'X-Requested-With': 'XMLHttpRequest',
        'Connection': 'keep-alive',
        'X-Anit-Forge-Token': 'None',
    }
    row = 1 if not red.get('row') else int(red.get('row'))
    now_page = 1 if not red.get('now_page') else int(red.get('now_page'))
    for i in range(now_page, max_page_number + 1):
        data = {
            'first': False,
            'pn': i,
            'kd': search_word,
        }

        resp = requests.post(
            'https://www.lagou.com/jobs/positionAjax.json?city=%s&needAddtionalResult=false' % city,
            headers=headers,
            cookies=cookie, data=data, verify=False)  # 请求接口
        resp_dict = resp.json()

        while not resp_dict.get('success'):
            print('请求频繁失败 等一会请求')
            time.sleep(25)

            url = 'https://www.lagou.com/jobs/list_%s/p-city_6?&cl=false&fromSearch=true&labelWords=&suginput=' % search_word
            resp = requests.get(url, headers=headers)  # 请求原网页

            resp_cookie = requests.utils.dict_from_cookiejar(resp.cookies)  # 获取cookies

            print(resp_cookie)
            if len(resp_cookie) <= 2:

                time.sleep(10)
                first_request(search_word)

            else:
                cookies.update(resp_cookie)

            resp = requests.post(
                'https://www.lagou.com/jobs/positionAjax.json?city=%s&needAddtionalResult=false' % city,
                headers=headers,
                cookies=cookie, data=data, verify=False)  # 请求接口
            resp_dict = resp.json()
            if resp_dict.get('success'):
                resp_cookie = requests.utils.dict_from_cookiejar(resp.cookies)  # 获取cookies
                cookie.update(resp_cookie)  # 更新接口的cookies
                row = 1 if not red.get('row') else int(red.get('row'))
                content_dict = resp_dict['content']
                positionResult = content_dict['positionResult']
                result = positionResult['result']
                showId = content_dict['showId']
                if result:
                    row = third_request(row, result, showId, search_word)
                    red.set('row', row)
                    break

        else:
            resp_cookie = requests.utils.dict_from_cookiejar(resp.cookies)  # 获取cookies
            cookie.update(resp_cookie)  # 更新接口的cookies
            row = 1 if not red.get('row') else int(red.get('row'))
            content_dict = resp_dict['content']
            positionResult = content_dict['positionResult']
            result = positionResult['result']
            showId = content_dict['showId']
            if result:
                row = third_request(row, result, showId, search_word)
                red.set('row', row)
        time.sleep(3)
    return row


def first_request(search_word):
    url = 'https://www.lagou.com/jobs/list_%s/p-city_6?&cl=false&fromSearch=true&labelWords=&suginput=' % search_word
    resp = requests.get(url, headers=headers)  # 请求原网页

    resp_cookie = requests.utils.dict_from_cookiejar(resp.cookies)  # 获取cookies

    print(resp_cookie)
    if len(resp_cookie) <= 2:

        time.sleep(10)
        first_request(search_word)

    else:
        cookies.update(resp_cookie)  # 更新接口的cookies
        max_page = etree.HTML(resp.text).xpath('//*[@id="order"]/li/div[4]/div[3]/span[2]')[0].text
        max_page_number = int(max_page)
        second_request(cookies, max_page_number, search_word)


def main():
    now_search = 0 if not red.get('now_search') else int(red.get('now_search'))
    for i in range(now_search, len(search_list)):
        red.set('now_search', i)
        first_request(search_list[i])
        time.sleep(10)


if __name__ == '__main__':
    main()
